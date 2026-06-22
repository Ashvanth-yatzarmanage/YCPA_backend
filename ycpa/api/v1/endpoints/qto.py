import io
import logging
from typing import Literal
from uuid import UUID

from fastapi import APIRouter, Query, status
from fastapi.responses import StreamingResponse

from ycpa.core.auth.dependencies import CurrentUser
from ycpa.core.database.dependencies import DatabaseSession
from ycpa.core.schemas.responses import SuccessResponse
from ycpa.schemas.requests.qto import QtoExtractRequest
from ycpa.schemas.responses.qto import QtoResultResponse
from ycpa.services.qto import QtoService

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/qto", tags=["QTO"])



@router.get(
    "/latest",
    response_model=SuccessResponse[QtoResultResponse],
    status_code=status.HTTP_200_OK,
    summary="Get latest QTO result for a project (used by Cost + Scheduling modules)",
)
async def get_latest_qto(
    current_user: CurrentUser,
    session:      DatabaseSession,
    project_id:   UUID  = Query(...),
    owner_type:   str   = Query(..., description="pim_project | aim_project"),
) -> SuccessResponse[QtoResultResponse]:

    from ycpa.core.exceptions import NotFoundException
    from ycpa.repositories.cde import CdeFileRepository

    repo = CdeFileRepository(session)

    files = await repo.get_by_owner(
        owner_type = owner_type,
        owner_id   = project_id,
        folder_id  = None,
    )
    ifc_files = [f for f in files if f.file_extension == 'ifc' and f.deleted_at is None]

    if not ifc_files:
        raise NotFoundException("No IFC files found in this project. Upload an IFC file in the CDE tab first.")

    latest = sorted(ifc_files, key=lambda f: f.created_at, reverse=True)[0]

    service = QtoService(session)
    result  = await service.extract(
        file_id      = latest.id,
        project_id   = project_id,
        owner_type   = owner_type,
        current_user = current_user,
    )
    return SuccessResponse(
        success = True,
        message = f"Latest QTO: {result.total_elements} elements from {result.filename}",
        data    = result,
    )



@router.post(
    "/extract",
    response_model=SuccessResponse[QtoResultResponse],
    status_code=status.HTTP_200_OK,
    summary="Extract quantities from an IFC file stored in CDE",
)
async def extract_qto(
    body:         QtoExtractRequest,
    current_user: CurrentUser,
    session:      DatabaseSession,
) -> SuccessResponse[QtoResultResponse]:
    service = QtoService(session)
    result  = await service.extract(
        file_id    = body.file_id,
        project_id = body.project_id,
        owner_type = body.owner_type,
        current_user = current_user,
    )
    return SuccessResponse(
        success = True,
        message = f"Extracted {result.total_elements} elements from {result.filename}",
        data    = result,
    )



@router.post(
    "/export",
    status_code=status.HTTP_200_OK,
    summary="Export QTO result as professional Excel (.xlsx) — 4 sheets",
)
async def export_qto(
    body:         QtoExtractRequest,
    current_user: CurrentUser,
    session:      DatabaseSession,
):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
    from openpyxl.utils import get_column_letter

    service = QtoService(session)
    result  = await service.extract(
        file_id      = body.file_id,
        project_id   = body.project_id,
        owner_type   = body.owner_type,
        current_user = current_user,
    )

    C_ACCENT    = "1E3A5F"
    C_ACCENT2   = "2563EB"
    C_LEVEL     = "1E3A5F"
    C_CAT       = "2D5FA6"
    C_FAM       = "D6E4F7"
    C_EL        = "FFFFFF"
    C_EL_ALT    = "F8FAFD"
    C_TOTAL     = "DBEAFE"
    C_BORDER    = "C7D9F0"

    def font(bold=False, color="1E293B", size=9, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border(all_sides=True, color=C_BORDER):
        s = Side(style="thin", color=color)
        n = Side(style=None)
        if all_sides:
            return Border(left=s, right=s, top=s, bottom=s)
        return Border(left=s, right=s, top=n, bottom=s)

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _c(ws, row, col, value, **kwargs):
        cell = ws.cell(row=row, column=col, value=value)
        if "font"      in kwargs: cell.font      = kwargs["font"]
        if "fill"      in kwargs: cell.fill      = kwargs["fill"]
        if "alignment" in kwargs: cell.alignment = kwargs["alignment"]
        if "border"    in kwargs: cell.border    = kwargs["border"]
        return cell

    def _num(v):
        """Round float for display."""
        return round(v, 4) if v is not None else None

    def _dash(v):
        return _num(v) if v is not None else "—"

    flat_rows = []
    outline_rows = []
    outline_counter = [0]

    for lv_idx, lv in enumerate(result.levels, start=1):
        outline_rows.append((lv.level, str(lv_idx), lv.level,
                             lv.count, None, None, None, []))
        for g_idx, grp in enumerate(lv.categories, start=1):
            outline_rows.append((lv.level, f"{lv_idx}.{g_idx}", grp.category,
                                 grp.count, grp.net_surface_area,
                                 grp.outer_surface_area, grp.net_volume, []))
            for f_idx, fam in enumerate(grp.families, start=1):
                fam_els = fam["elements"] if isinstance(fam, dict) else fam.elements
                fam_count = fam["count"] if isinstance(fam, dict) else fam.count
                fam_name  = fam["name"]  if isinstance(fam, dict) else fam.name
                fam_nsa   = fam["net_surface_area"]   if isinstance(fam, dict) else fam.net_surface_area
                fam_osa   = fam["outer_surface_area"] if isinstance(fam, dict) else fam.outer_surface_area
                fam_vol   = fam["net_volume"]         if isinstance(fam, dict) else fam.net_volume
                outline_rows.append((lv.level,
                                     f"{lv_idx}.{g_idx}.{f_idx}", fam_name,
                                     fam_count, fam_nsa, fam_osa, fam_vol, fam_els))
                for el in fam_els:
                    flat_rows.append((el.internal_id, el.global_id,
                                      lv.level, grp.category, fam_name,
                                      el.length, el.width, el.height,
                                      el.net_surface_area, el.net_volume,
                                      el.outer_surface_area))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Styled_Report")

    ws1.merge_cells("A1:F1")
    c = ws1["A1"]
    c.value     = "QUANTITY TAKE-OFF REPORT"
    c.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    c.fill      = fill(C_ACCENT)
    c.alignment = align("left")
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:F2")
    c = ws1["A2"]
    c.value     = f"File: {result.filename}    |    Extracted: {result.extracted_at.strftime('%Y-%m-%d %H:%M UTC')}    |    Total: {result.total_elements} elements"
    c.font      = font(size=9, color="64748B", italic=True)
    c.fill      = fill("EFF6FF")
    c.alignment = align("left")
    ws1.row_dimensions[2].height = 16

    ws1.append([])

    current_row = 4

    COLS_HDR = ["Building Level", "Category", "Family / Type",
                "Count", "Net Surface Area (m²)", "Net Volume (m³)"]
    for ci, h in enumerate(COLS_HDR, 1):
        _c(ws1, current_row, ci, h,
           font=font(bold=True, color="FFFFFF", size=9),
           fill=fill(C_ACCENT2),
           alignment=align("center"),
           border=border())
    ws1.row_dimensions[current_row].height = 18
    current_row += 1

    for lv in result.levels:
        # Level row
        ws1.merge_cells(start_row=current_row, start_column=1,
                        end_row=current_row, end_column=6)
        _c(ws1, current_row, 1, lv.level,
           font=font(bold=True, color="FFFFFF", size=10),
           fill=fill(C_LEVEL),
           alignment=align("left"),
           border=border())
        ws1.row_dimensions[current_row].height = 18
        current_row += 1

        for grp in lv.categories:
            _c(ws1, current_row, 1, "",
               fill=fill(C_CAT), border=border())
            _c(ws1, current_row, 2, grp.category,
               font=font(bold=True, color="FFFFFF", size=9),
               fill=fill(C_CAT), alignment=align("left"), border=border())
            for ci in range(3, 7):
                _c(ws1, current_row, ci, "",
                   fill=fill(C_CAT), border=border())
            ws1.row_dimensions[current_row].height = 16
            current_row += 1

            for fam in grp.families:
                fam_name = fam["name"]  if isinstance(fam, dict) else fam.name
                fam_count= fam["count"] if isinstance(fam, dict) else fam.count
                fam_nsa  = fam["net_surface_area"]   if isinstance(fam, dict) else fam.net_surface_area
                fam_vol  = fam["net_volume"]         if isinstance(fam, dict) else fam.net_volume
                fam_els  = fam["elements"] if isinstance(fam, dict) else fam.elements

                _c(ws1, current_row, 1, lv.level,
                   font=font(size=9, color="475569"),
                   fill=fill(C_FAM), alignment=align("left"), border=border())
                _c(ws1, current_row, 2, grp.category,
                   font=font(size=9, color="475569"),
                   fill=fill(C_FAM), alignment=align("left"), border=border())
                _c(ws1, current_row, 3, fam_name,
                   font=font(bold=True, size=9, color="1E3A5F"),
                   fill=fill(C_FAM), alignment=align("left"), border=border())
                _c(ws1, current_row, 4, fam_count,
                   font=font(bold=True, size=9),
                   fill=fill(C_FAM), alignment=align("center"), border=border())
                _c(ws1, current_row, 5, _dash(fam_nsa),
                   font=font(bold=True, size=9),
                   fill=fill(C_FAM), alignment=align("right"), border=border())
                _c(ws1, current_row, 6, _dash(fam_vol),
                   font=font(bold=True, size=9),
                   fill=fill(C_FAM), alignment=align("right"), border=border())
                ws1.row_dimensions[current_row].height = 16
                current_row += 1

                for ei, el in enumerate(fam_els):
                    bg = C_EL if ei % 2 == 0 else C_EL_ALT
                    _c(ws1, current_row, 1, "",
                       fill=fill(bg), border=border())
                    _c(ws1, current_row, 2, "",
                       fill=fill(bg), border=border())
                    _c(ws1, current_row, 3, f"  [{el.internal_id}]",
                       font=font(size=8, color="64748B"),
                       fill=fill(bg), alignment=align("left"), border=border())
                    _c(ws1, current_row, 4, 1,
                       font=font(size=8), fill=fill(bg),
                       alignment=align("center"), border=border())
                    _c(ws1, current_row, 5, _dash(el.net_surface_area),
                       font=font(size=8, color="334155"),
                       fill=fill(bg), alignment=align("right"), border=border())
                    _c(ws1, current_row, 6, _dash(el.net_volume),
                       font=font(size=8, color="334155"),
                       fill=fill(bg), alignment=align("right"), border=border())
                    ws1.row_dimensions[current_row].height = 14
                    current_row += 1

        lv_nsa = sum((g.net_surface_area or 0) for g in lv.categories)
        lv_vol = sum((g.net_volume       or 0) for g in lv.categories)
        for ci, val in enumerate([f"  TOTAL — {lv.level}", "", "", lv.count,
                                   round(lv_nsa, 4), round(lv_vol, 4)], 1):
            _c(ws1, current_row, ci, val,
               font=font(bold=True, size=9, color=C_ACCENT2),
               fill=fill(C_TOTAL),
               alignment=align("right" if ci >= 4 else "left"),
               border=border())
        ws1.row_dimensions[current_row].height = 16
        current_row += 2  # blank line between levels

    grand_nsa = sum((g.net_surface_area or 0) for lv in result.levels for g in lv.categories)
    grand_vol = sum((g.net_volume       or 0) for lv in result.levels for g in lv.categories)
    for ci, val in enumerate(["GRAND TOTAL", "", "", result.total_elements,
                               round(grand_nsa, 4), round(grand_vol, 4)], 1):
        _c(ws1, current_row, ci, val,
           font=font(bold=True, size=10, color="FFFFFF"),
           fill=fill(C_ACCENT),
           alignment=align("right" if ci >= 4 else "left"),
           border=border())
    ws1.row_dimensions[current_row].height = 20

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 38
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 22
    ws1.column_dimensions["F"].width = 18
    ws1.freeze_panes = "A5"

    ws2 = wb.create_sheet("Flat_Table")

    FLAT_COLS = [
        "Element ID (Internal)", "Element ID (Global)", "Building Level",
        "Category", "Family",
        "Count (Unit)", "Count (Value)",
        "BB Length (Unit)", "BB Length (Value)",
        "BB Width (Unit)",  "BB Width (Value)",
        "BB Height (Unit)", "BB Height (Value)",
        "NetSurfaceArea (Unit)", "NetSurfaceArea (Value)",
        "NetVolume (Unit)", "NetVolume (Value)",
        "OuterSurfaceArea (Unit)", "OuterSurfaceArea (Value)",
    ]
    for ci, h in enumerate(FLAT_COLS, 1):
        _c(ws2, 1, ci, h,
           font=font(bold=True, color="FFFFFF", size=9),
           fill=fill(C_ACCENT),
           alignment=align("center"),
           border=border())
    ws2.row_dimensions[1].height = 18

    for ri, row in enumerate(flat_rows, start=2):
        (int_id, glob_id, level, cat, fam,
         length, width, height, nsa, vol, osa) = row
        bg = C_EL if ri % 2 == 0 else C_EL_ALT
        vals = [
            int_id, glob_id, level, cat, fam,
            "", 1,
            "m", _num(length),
            "m", _num(width),
            "m", _num(height),
            "m²", _num(nsa),
            "m³", _num(vol),
            "m²", _num(osa),
        ]
        for ci, v in enumerate(vals, 1):
            _c(ws2, ri, ci, v,
               font=font(size=8),
               fill=fill(bg),
               alignment=align("right" if ci >= 7 else "left"),
               border=border())
        ws2.row_dimensions[ri].height = 14

    for ci, w in enumerate([14, 36, 20, 14, 36,
                              8,  8,  8, 14,  8, 14,
                              8, 14,  8, 18,  8, 14,
                              8, 18], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Breakdown_Structure")

    BS_COLS = [
        "Outline Level",
        f"Project — {result.filename} ({result.total_elements} Elements)",
        "Count (Unit)", "Count (Value)",
        "NetSurfaceArea (Unit)", "NetSurfaceArea (Value)",
        "NetVolume (Unit)", "NetVolume (Value)",
        "OuterSurfaceArea (Unit)", "OuterSurfaceArea (Value)",
    ]
    for ci, h in enumerate(BS_COLS, 1):
        _c(ws3, 1, ci, h,
           font=font(bold=True, color="FFFFFF", size=9),
           fill=fill(C_ACCENT),
           alignment=align("center"),
           border=border())
    ws3.row_dimensions[1].height = 18

    ri = 2
    for lv_idx, lv in enumerate(result.levels, 1):
        lv_nsa = sum((g.net_surface_area or 0) for g in lv.categories)
        lv_osa = sum((g.outer_surface_area or 0) for g in lv.categories)
        lv_vol = sum((g.net_volume or 0) for g in lv.categories)

        row_data = [str(lv_idx), lv.level, "", lv.count,
                    "m²", _num(lv_nsa), "m³", _num(lv_vol), "", ""]
        for ci, v in enumerate(row_data, 1):
            _c(ws3, ri, ci, v,
               font=font(bold=True, color="FFFFFF", size=9),
               fill=fill(C_LEVEL),
               alignment=align("right" if ci >= 4 else "left"),
               border=border())
        ws3.row_dimensions[ri].height = 16
        ri += 1

        for g_idx, grp in enumerate(lv.categories, 1):
            row_data = [f"{lv_idx}.{g_idx}", grp.category, "", grp.count,
                        "m²", _num(grp.net_surface_area),
                        "m³", _num(grp.net_volume),
                        "m²", _num(grp.outer_surface_area)]
            for ci, v in enumerate(row_data, 1):
                _c(ws3, ri, ci, v,
                   font=font(bold=True, color="FFFFFF", size=9),
                   fill=fill(C_CAT),
                   alignment=align("right" if ci >= 4 else "left"),
                   border=border())
            ws3.row_dimensions[ri].height = 15
            ri += 1

            for f_idx, fam in enumerate(grp.families, 1):
                fam_name  = fam["name"]               if isinstance(fam, dict) else fam.name
                fam_count = fam["count"]               if isinstance(fam, dict) else fam.count
                fam_nsa   = fam["net_surface_area"]    if isinstance(fam, dict) else fam.net_surface_area
                fam_osa   = fam["outer_surface_area"]  if isinstance(fam, dict) else fam.outer_surface_area
                fam_vol   = fam["net_volume"]          if isinstance(fam, dict) else fam.net_volume

                row_data = [f"{lv_idx}.{g_idx}.{f_idx}", fam_name, "", fam_count,
                            "m²", _num(fam_nsa),
                            "m³", _num(fam_vol),
                            "m²", _num(fam_osa)]
                bg = C_FAM
                for ci, v in enumerate(row_data, 1):
                    _c(ws3, ri, ci, v,
                       font=font(size=9, color="1E3A5F"),
                       fill=fill(bg),
                       alignment=align("right" if ci >= 4 else "left"),
                       border=border())
                ws3.row_dimensions[ri].height = 14
                ri += 1

    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 40
    for ci in range(3, 11):
        ws3.column_dimensions[get_column_letter(ci)].width = 14
    ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet("Breakdown_Structure_Elements")

    for ci, h in enumerate(BS_COLS, 1):
        _c(ws4, 1, ci, h,
           font=font(bold=True, color="FFFFFF", size=9),
           fill=fill(C_ACCENT),
           alignment=align("center"),
           border=border())
    ws4.row_dimensions[1].height = 18

    ri = 2
    for lv_idx, lv in enumerate(result.levels, 1):
        lv_nsa = sum((g.net_surface_area or 0) for g in lv.categories)
        lv_vol = sum((g.net_volume or 0) for g in lv.categories)

        for ci, v in enumerate([str(lv_idx), lv.level, "", lv.count,
                                  "m²", _num(lv_nsa), "m³", _num(lv_vol), "", ""], 1):
            _c(ws4, ri, ci, v,
               font=font(bold=True, color="FFFFFF", size=9),
               fill=fill(C_LEVEL),
               alignment=align("right" if ci >= 4 else "left"),
               border=border())
        ws4.row_dimensions[ri].height = 16; ri += 1

        for g_idx, grp in enumerate(lv.categories, 1):
            for ci, v in enumerate([f"{lv_idx}.{g_idx}", grp.category, "", grp.count,
                                     "m²", _num(grp.net_surface_area),
                                     "m³", _num(grp.net_volume),
                                     "m²", _num(grp.outer_surface_area)], 1):
                _c(ws4, ri, ci, v,
                   font=font(bold=True, color="FFFFFF", size=9),
                   fill=fill(C_CAT),
                   alignment=align("right" if ci >= 4 else "left"),
                   border=border())
            ws4.row_dimensions[ri].height = 15; ri += 1

            for f_idx, fam in enumerate(grp.families, 1):
                fam_name  = fam["name"]              if isinstance(fam, dict) else fam.name
                fam_count = fam["count"]              if isinstance(fam, dict) else fam.count
                fam_nsa   = fam["net_surface_area"]   if isinstance(fam, dict) else fam.net_surface_area
                fam_osa   = fam["outer_surface_area"] if isinstance(fam, dict) else fam.outer_surface_area
                fam_vol   = fam["net_volume"]         if isinstance(fam, dict) else fam.net_volume
                fam_els   = fam["elements"]           if isinstance(fam, dict) else fam.elements

                for ci, v in enumerate([f"{lv_idx}.{g_idx}.{f_idx}", fam_name, "", fam_count,
                                         "m²", _num(fam_nsa),
                                         "m³", _num(fam_vol),
                                         "m²", _num(fam_osa)], 1):
                    _c(ws4, ri, ci, v,
                       font=font(bold=True, size=9, color="1E3A5F"),
                       fill=fill(C_FAM),
                       alignment=align("right" if ci >= 4 else "left"),
                       border=border())
                ws4.row_dimensions[ri].height = 14; ri += 1

                for el_idx, el in enumerate(fam_els, 1):
                    bg = C_EL if el_idx % 2 == 0 else C_EL_ALT
                    outline = f"{lv_idx}.{g_idx}.{f_idx}.{el_idx}"
                    for ci, v in enumerate([outline, f"[{el.internal_id}]", "", 1,
                                             "m²", _num(el.net_surface_area),
                                             "m³", _num(el.net_volume),
                                             "m²", _num(el.outer_surface_area)], 1):
                        _c(ws4, ri, ci, v,
                           font=font(size=8, color="64748B"),
                           fill=fill(bg),
                           alignment=align("right" if ci >= 4 else "left"),
                           border=border())
                    ws4.row_dimensions[ri].height = 13; ri += 1

    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 40
    for ci in range(3, 11):
        ws4.column_dimensions[get_column_letter(ci)].width = 14
    ws4.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = result.filename.replace(".ifc", "").replace(" ", "_")
    date_str  = result.extracted_at.strftime("%Y-%m-%d")

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="QTO_{safe_name}_{date_str}.xlsx"'
        },
    )